#######################################################################################
################################ 초기 세팅 방법 #########################################
### 1. https://www.python.org/downloads/ 설치
### 2. 해당 exe 설치 진행 시 아래에 administrator 실행 및 PATH 추가 선택
### 3. 동일 파일에 추가 되어 있는 파일 설치 pip install "file.whl" 실행
### 4. .py 파일을 notepad로 실행 시켜 pdf 파일 및 excel 파일 위치 설정
### 5. .py 파일 실행
#################################### 끝 ################################################
########################################################################################


import os
import fitz  # PyMuPDF
import re
import openpyxl
from datetime import date, timedelta, datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import column_index_from_string
import random

tt = ["LJH_Sexy_Guy", "LJH_Macho_Guy", "LJH_Best_Guy", "Sexy_Master", "Dance_King", "Future_Leader"]
The_King = random.choice(tt)
# The_King = "LJH_Sexy_Guy"


## 월을 숫자 → 영문으로 변환
def mon(text):
    """월(MM)을 영문 월(JAN, FEB 등)로 변환"""
    months = {
        "01": "JAN", "02": "FEB", "03": "MAR", "04": "APR", "05": "MAY", "06": "JUN",
        "07": "JUL", "08": "AUG", "09": "SEP", "10": "OCT", "11": "NOV", "12": "DEC"
    }
    return months.get(text, "")

## PO Number page가 달라 못받을 경우 H6에 있는 값 받아오기
def fill_missing_pno(ws):
    """H열(pno)에 값이 없는 경우 기본값 채워 넣기"""
    column_index = column_index_from_string("H")  # H = 8
    first_row = 7  # 데이터는 H6부터 시작

    for row in range(first_row, ws.max_row + 1):
        cell = ws.cell(row=row, column=column_index)
        if cell.value is None or str(cell.value).strip() == "":
            cell.value = "=H6"

## 월을 영문 → 숫자로 변환
def rev_mon(text):
    """월(MM)을 영문 월(JAN, FEB 등)로 변환"""
    months = {
        "JAN": "01", "FEB": "02", "MAR": "03", "APR": "04", "MAY": "05", "JUN": "06",
        "JUL": "07", "AUG": "08", "SEP": "09", "OCT": "10", "NOV": "11", "DEC": "12"
    }
    return months.get(text, "")

## 월/년 만 있을 경우 해당 달에 해당하는 첫번째 월요일 받아오기
def first_monday(the_month, the_year):
    month = int(the_month)
    year = int(the_year)
    first_day = date(year, month, 1)
    days_until_Monday = (0 - first_day.weekday()) % 7
    the_Monday = first_day + timedelta(days=days_until_Monday)
    return the_Monday.day

def get_materialcode(text):
    """월(MM)을 영문 월(JAN, FEB 등)로 변환"""
    material = {
        "9120491 ASA LI941 F94484 (LG)": "9120491 ASA LI941 F94484 (LG)", 
        "9122188 ASA LI 941V NEGRO 9B9 (LG)": "9122188 ASA LI 941V NEGRO 9B9 (LG)", 
        "LG LI941V_V94841_ASA": "LG LI941V_V94841_ASA", 
        "5335630000": "5335630000", 
        "ABS LG ER400 M95007 schwarz": "ABS LG ER400 M95007 schwarz", 
        "0075A00054100GR": "0075A00054100GR",
        "LG ASA LI941-V - 94841 (VW9B9) (SILO)": "LG ASA LI941-V - 94841 (VW9B9) (SILO)", 
        "LG ASA LI941-F - 94841 (VW9B9) AEB": "LG ASA LI941-F - 94841 (VW9B9) AEB", 
        "LG ASA LI941 - V94841 (VW9B9) BigBag": "LG ASA LI941 - V94841 (VW9B9) BigBag", 
        "High gloss ASA: LI941F Piano Black (F94484)": "High gloss ASA: LI941F Piano Black (F94484)",
        "ABS XR 401 BK 9001": "ABS XR 401 BK 9001", 
        "High gloss ASA LI941-F94484 (Piano Black)": "High gloss ASA LI941-F94484 (Piano Black)",
        "30022028 LG LI941-F 94484 PIANO BLACK": "30022028 LG LI941-F 94484 PIANO BLACK",
        "30021896 LG LI941V 94841": "30021896 LG LI941V 94841",
        "30022062 ABS XR 410 NATUR": "30022062 ABS XR 410 NATUR",
        "ASA LI941V": "ASA LI941V",
        "ASA LI941F-94841": "ASA LI941F-94841",
        "LG LI941 F 94484 PIANO BLACK": "LG LI941 F 94484 PIANO BLACK",
        "ABS XR 410 NATUR": "ABS XR 410 NATUR", # 중복됨.
        "ABS ER400-M95007": "ABS ER400-M95007",
        "ASALI941-F94841 (9B9)": "ASALI941-F94841 (9B9)",
        "ABS 950kg XR410 9529": "ABS 950kg XR410 9529",
        "ABS XR410 naturverpackt in Octabin": "ABS XR410 naturverpackt in Octabin",
        "ASA LI941 F94484 (LG)": "ASA LI941 F94484 (LG)",
        "ABS ER400 M95007 schwarz": "ABS ER400 M95007 schwarz",
        "ABS ER400 M97005 NEGRO": "ABS ER400 M97005 NEGRO",
        "None": "Material Code 모름"
    }
    return material.get(text, "") 

## PO Number page가 달라 못받을 경우 H6에 있는 값 받아오기
def fill_missing_pno(ws):
    """H열(pno)에 값이 없는 경우 기본값 채워 넣기"""
    column_index = column_index_from_string("H")  # H = 8
    first_row = 7  # 데이터는 H6부터 시작

    for row in range(first_row, ws.max_row + 1):
        cell = ws.cell(row=row, column=column_index)
        if cell.value is None or str(cell.value).strip() == "":
            cell.value = "=H6"

## (H열) 마지막 줄 찾기
def get_last_filled_row(ws, column_letter="H"):
    column_index = column_index_from_string(column_letter)  # 'H' => 8
    last_row = 6  # 최소 4행 보장 (H열 4행까지는 무조건 값이 있음)
    
    for row in range(6, ws.max_row + 1):  # 5행부터 검색
        if ws.cell(row=row, column=column_index).value:  # 값이 있는 마지막 행 찾기
            last_row = row
    
    return last_row  # 최소 4행 보장

## 시트별 첫 두 행에 회사명 및 Material Code 추가
def first_two_lows(name, ws, material):
    ws.merge_cells('B1:H1')
    ws.row_dimensions[1].height = 30
    ws['B1'] = name
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B1'].font = Font(size=15, bold=True)

    ws.merge_cells('B2:H2')
    # ws.row_dimensions[2].height = 30
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B2'].font = Font(bold=True)
    ws['B2'] = material

## 시트별 디자인 구성성
def get_or_create_sheet(wb, sheet_name):
    """Excel 시트를 가져오거나 없으면 새로 생성"""

    sheet_name = sheet_name[:31]
    
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    
    ws = wb.create_sheet(sheet_name)

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 15
    ws.column_dimensions['H'].width = 35

    # 첫 번째 행: 파이썬 파일 돌린 시간
    ws.merge_cells('B1:H1')
    ws['B1'] = datetime.now()
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B1'].font = Font(bold=True)

    # 두 번째 행: 뭐 적어야 되는지 모름
    ws.merge_cells('B2:H2')
    # ws.row_dimensions[2].height = 30
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B2'].font = Font(bold=True)
    ws['B2'] = "JUN HYEOK LEE!!!!!! The Master of Logistics"

    # 세 번째 행: 인바운드 & 아웃 바운드
    ws.merge_cells('B3:D3')
    Green_Fill = PatternFill(start_color="00FF00", end_color="FF9999", fill_type="mediumGray")
    ws['B3'].fill = Green_Fill
    ws['B3'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B3'].font = Font(bold=True)
    ws['B3'] = "Inbound"

    ws.merge_cells('F3:H3')
    Red_Fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="mediumGray")
    ws['F3'].fill = Red_Fill
    ws['F3'].alignment = Alignment(horizontal='center', vertical='center')
    ws['F3'].font = Font(bold=True)
    ws['F3'] = "Outbound"
    
    # 네 번째 행: 열 제목(데이터 종류)
    ws.append(["Month", "PO No", "Date", "QTY (MT)", "On Stock", "QTY (MT)", "Date", "PO No", "Filename"])
    for col in range(1,9):
        cell = ws.cell(row=4, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')        
    
    # 다섯 번째 행: 빈 값
    ws.append([""] * 8)

    return ws

### 색상 추가 부분!! 필요 없을 경우 삭제
def apply_conditional_formatting(ws, max_row):
    """D3 값을 기준으로 D4:D1000 범위에 조건부 서식 적용"""

    # if max_row is None or max_row < 7:
    #     max_row = 100
    
    # 색상 정의
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # 초록색
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 노란색
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # 주황색
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # 빨간색

    # 범위 정의
    data_range = f"E6:E{max_row}"

    # thick_border = Border(
    #         left = Side(style="thick"),
    #         right = Side(style="thick"),
    #         top = Side(style="thick"),
    #         bottom = Side(style="thick")
    #     )
    
    # for row in range(6, max_row+1):
    #     cell = ws.cell(row=row, column=5)
    #     cell.border = thick_border

    # 조건부 서식 추가 (D3을 기준으로 계산)
    ws.conditional_formatting.add(
        data_range,
        FormulaRule(formula=["E6>=E$5*0.6"], stopIfTrue=True, fill=green_fill)  # 60% 이상 초록색
    )
    ws.conditional_formatting.add(
        data_range,
        FormulaRule(formula=["AND(E6>=E$5*0.4, E6<E$5*0.6)"], stopIfTrue=True, fill=yellow_fill)  # 40% 이상 노란색
    )
    ws.conditional_formatting.add(
        data_range,
        FormulaRule(formula=["AND(E6>0, E6<E$5*0.4)"], stopIfTrue=True, fill=orange_fill)  # 40% 미만 주황색
    )
    ws.conditional_formatting.add(
        data_range,
        FormulaRule(formula=["E6<0"], stopIfTrue=True, fill=red_fill)  # 음수(마이너스 값) 빨간색
    )
    # 0인 경우 색상 없음 (기본값 유지)

    # 셀 테두리 변경
    thick = Side(style="thick")
    thin = Side(style="thin")
    double = Side(style="double")
    medium = Side(style="medium")

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.border = Border(right=double)

    for cell in ws[5]:
        cell.border = Border(top=double)

    ws['A5'].border = Border(right=double, top=double)

    first_row = 4
    last_row = max_row
    column = 5
    for row in range(first_row, last_row+1):
        cell = ws.cell(row=row, column=column)
        cell2 = ws.cell(row=row, column=1)

        if row == first_row:
            cell.border = Border(top=thick, left=thick, right=thick, bottom=None)
            # cell2.border = Border(top=medium, left=medium, right=medium, bottom=None)
        elif row == last_row:
            cell.border = Border(top=None, left=thick, right=thick, bottom=thick)
            # cell2.border = Border(top=None, left=medium, right=medium, bottom=medium)
        else:
            cell.border = Border(top=None, bottom=None, left=thick, right=thick)
            # cell2.border = Border(top=None, bottom=None, left=medium, right=medium)



######################################################
################ 회사별 함수 분류 ######################
######################################################

# 기준: "SMP Ibérica S.L.U."
def process_smp_iberica(text, filename, ws):
    """SMP IBERICA S.L.U. /PALENCIA 문서를 처리하는 함수"""
    print(f"📄 Processing file: {filename}")

    lines = text.replace(",", "").replace(".", "").split("\n")
    cleaned_list = [item.strip() for item in lines if item.strip()]
    written_date = ""
    written_month = ""
    pno = ""
    i = 0
    merge_lines = []

    while i < len(cleaned_list):
        if cleaned_list[i].strip().startswith("W ") and i + 1 < len(cleaned_list):
            merge_lines.append(f"{cleaned_list[i]} {cleaned_list[i+1]}")
            i+=2
        elif cleaned_list[i].strip().endswith("number/date") and i + 1 < len(cleaned_list):
            merge_lines.append(f"{cleaned_list[i]} {cleaned_list[i+1]}")
            i+=2
        else:
            merge_lines.append(str(cleaned_list[i]))
            i += 1
    
    for line in merge_lines:
        extracted_texts = re.split(r'\s+', line)
        extracted_text = [item for item in extracted_texts if item.strip()]
        material = get_materialcode("9120491 ASA LI941 F94484 (LG)")

        if "number/date" in line:
            if len(extracted_text) >= 4:
                pno = extracted_text[3]
                first_row = 6
                last_row = ws.max_row
                column = 8
                for row in range(first_row, last_row+1):
                    cell = ws.cell(row=row, column=column)
                    cell.value = pno    

        if line.startswith("W ") or line.startswith("D "):
            if len(extracted_text) >= 3:
                quantity = extracted_text[2]
                if quantity.isdigit() and int(quantity) > 0:
                    qty = int(quantity)/100000
                    ### 시작
                    written_date = f"{extracted_text[1][:2]}-{extracted_text[1][2:4]}-{extracted_text[1][4:]}"
                    written_month = f"{mon(extracted_text[1][2:4])}-{extracted_text[1][6:]}"
                    # print(f"[SMP Ibérica] 데이터 추가: {quantity}")
                    ws.append([written_month, The_King, The_King, The_King, "On Stock", qty, written_date, pno, filename])
                    first_two_lows("SMP IBERICA S.L.U. /PALENCIA", ws, material)

# 기준: "Samvardhana Motherson Peguform"
def process_samvardhanaPeguform(text, filename, ws):
    """Samvardhana Motherson 문서를 처리하는 함수"""
    print(f"📄 Processing file: {filename}")

    lines = text.replace(",", "").replace(".", "").split("\n")
    cleaned_list = [item.strip() for item in lines if item.strip()]
    written_date = ""
    written_month = ""
    pno = ""
    i = 0
    merge_lines = []

    while i < len(cleaned_list):
        if cleaned_list[i].strip().endswith("number/date") and i + 1 < len(cleaned_list):
            merge_lines.append(f"{cleaned_list[i]} {cleaned_list[i+1]}")
            i+=2
        else:
            merge_lines.append(str(cleaned_list[i]))
            i += 1
    
    for line in merge_lines:
        extracted_texts = re.split(r'\s+', line)
        extracted_text = [item for item in extracted_texts if item.strip()]
        material = get_materialcode("9122188 ASA LI 941V NEGRO 9B9 (LG)")
        
        if "number/date" in line:
            if len(extracted_text) >= 4:
                pno = extracted_text[3]
                first_row = 6
                last_row = ws.max_row
                column = 8
                for row in range(first_row, last_row+1):
                    cell = ws.cell(row=row, column=column)
                    cell.value = pno           
        
        if line.startswith("W ") or line.startswith("D "):
            if len(extracted_text) >= 4:
                quantity = extracted_text[3]
                if quantity.isdigit() and int(quantity) > 0:
                    qty = float(quantity)/100000
                    written_date = f"{extracted_text[2][:2]}-{extracted_text[2][2:4]}-{extracted_text[2][4:]}"
                    written_month = f"{mon(extracted_text[2][2:4])}-{extracted_text[2][6:]}"
                    ws.append([written_month, The_King, The_King, The_King, "On Stock", qty, written_date, pno, filename])
                    first_two_lows("Samvardhana Motherson Peguform", ws, material)
                    # print(f"[Samvardhana Motherson] 데이터 추가: {quantity}")

# 기준: "Samvardhana Motherson Innovative"
def process_samvardhanaInnovative(text, filename, ws):
    """Samvardhana Motherson Innovative 문서를 처리하는 함수"""
    print(f"📄 Processing file: {filename}")

    lines = text.replace(",", "").replace(".", "").split("\n")
    cleaned_list = [item.strip() for item in lines if item.strip()]
    written_date = ""
    written_month = ""
    pno = ""
    i = 0
    merge_lines = []
    while i < len(cleaned_list):
        if cleaned_list[i].strip().startswith("D") and i + 1 < len(cleaned_list):           
            merge_lines.append(f"{cleaned_list[i]} {cleaned_list[i+1]} {cleaned_list[i+2]} {cleaned_list[i+3]} {cleaned_list[i+4]}")
            i+=5
        elif cleaned_list[i].strip().startswith("Sch Agr No") and i + 1 < len(cleaned_list):
            merge_lines.append(f"{cleaned_list[i-1]} {cleaned_list[i]}")
            i+=1
        else:
            merge_lines.append(str(cleaned_list[i]))
            i += 1
    
    for line in merge_lines:
        extracted_texts = re.split(r'\s+', line)
        extracted_text = [item for item in extracted_texts if item.strip()]
        material = get_materialcode("LG LI941V_V94841_ASA")
        if "Sch Agr No" in line:
            if len(extracted_text) >= 4:
                pno = extracted_text[0]
                first_row = 6
                last_row = ws.max_row
                column = 8
                for row in range(first_row, last_row+1):
                    cell = ws.cell(row=row, column=column)
                    cell.value = pno    
        
        if line.startswith("W ") or line.startswith("D"):
            if "Date" not in line:
                if len(extracted_text) >= 5:
                    quantity = extracted_text[2]
                    if quantity.isdigit() and int(quantity) > 0:
                        qty = float(quantity)/100000
                        written_date = f"{extracted_text[1][:2]}-{extracted_text[1][2:4]}-{extracted_text[1][4:]}"
                        written_month = f"{mon(extracted_text[1][2:4])}-{extracted_text[1][6:]}"
                        if written_date and pno:
                            # print(f"✅ 저장된 정보 기반 데이터 추가: {qty}, {written_date}, {pno}")
                            ws.append([written_month, The_King, The_King, The_King, "On Stock", qty, written_date, pno, filename]) 
                            first_two_lows("Samvardhana Motherson Innovative", ws, material)
           
# 기준: "PLASTICOS ABC SPAIN"
def process_PLASTICOS(text, filename, ws):
    """Plasticos abc Spain 문서를 처리하는 함수"""
    print(f"📄 Processing file: {filename}")

    lines = text.replace(",", "").replace(".", "").split("\n")
    cleaned_list = [item.strip() for item in lines if item.strip()]
    written_date = ""
    written_month = ""
    pno = ""
    material = get_materialcode("ABS ER400 M97005 NEGRO")

    for line in cleaned_list:
        extracted_texts = re.split(r'\s+', line)
        extracted_text = [item for item in extracted_texts if item.strip()]
        # print(f"@@@@@@@@@@@@@@@line {line}")
        # print(f"@@@@@@@@@@@@@@@extracted_text {extracted_text}")

        if "P/O" in line and "#:" in line:
            if int(extracted_text[7]) > 0:
                pno = extracted_text[7]
                first_row = 6
                last_row = ws.max_row
                column = 8                
                for row in range(first_row, last_row+1):
                    cell = ws.cell(row=row, column=column)
                    cell.value = pno
        
        if "Week" in line and len(extracted_text) > 5:
            quantity = extracted_text[3]
            if quantity.isdigit() and int(quantity) > 0:
                qty = float(quantity)/100000
                for_date = extracted_text[0].split("/")
                if for_date[0] != "RELEASES":
                    if len((for_date)[0]) < 2:
                        for_date[0] = f"0{for_date[0]}"
                    written_date = f"{for_date[0]}-{for_date[1]}-{for_date[2]}"
                    written_month = f"{mon(for_date[1])}-{for_date[2]}"
                    if written_date and pno:
                        # print(f"✅ 저장된 정보 기반 데이터 추가: {qty}, {written_date}, {pno}")
                        ws.append([written_month, The_King, The_King, The_King, "On Stock", qty, written_date, pno, filename]) 
                    # else:
                    #     print("❌ 날짜나 PNO 정보가 누락됨 → 건너뜀")
            # else:
            #     print("숫자로 변환 불가")        
            
        if "Month" in line and len(extracted_text) > 5:
            quantity = extracted_text[3]
            if quantity.isdigit() and int(quantity) > 0:
                qty = float(quantity)/100000
                for_date = extracted_text[0].split("/")
                if len((for_date)[0]) < 2:
                    for_date[0] = f"0{for_date[0]}"
                written_date = f"{for_date[0]}-{for_date[1]}-{for_date[2]}"
                written_month = f"{mon(for_date[1])}-{for_date[2]}"
                if written_date and pno:
                    # print(f"✅ 저장된 정보 기반 데이터 추가: {qty}, {written_date}, {pno}")
                    ws.append([written_month, The_King, The_King, The_King, "On Stock", qty, written_date, pno, filename]) 
                # else:
                #     print("❌ 날짜나 PNO 정보가 누락됨 → 건너뜀")
            # else:
            #     print("숫자로 변환 불가")

        # date1 to date2 Floating Forecast 나왔을 경우 date2를 표시. >> for_date = extracted_text[2].split("/")
        if "Floating" in line and len(extracted_text) > 8:
            if "Forecast" in line:
                quantity = extracted_text[5]
                if quantity.isdigit() and int(quantity) > 0:
                    qty = float(quantity)/100000
                    for_date = extracted_text[2].split("/")
                    if len((for_date)[0]) < 2:
                        for_date[0] = f"0{for_date[0]}"
                    written_date = f"{for_date[0]}-{for_date[1]}-{for_date[2]}"
                    written_month = f"{mon(for_date[1])}-{for_date[2]}"
                    if written_date and pno:
                        # print(f"✅ 저장된 정보 기반 데이터 추가: {qty}, {written_date}, {pno}")
                        ws.append([written_month, The_King, The_King, The_King, "On Stock", qty, written_date, pno, filename]) 
                    else:
                        print("❌ 날짜나 PNO 정보가 누락됨 → 건너뜀")

        first_two_lows("PLASTICOS ABC SPAIN", ws, material)
                    # ws.append([written_month, The_King, datetime.now().strftime("%Y-%m-%d"), The_King, "On Stock", qty, written_date, pno])
                    # print(f"[Samvardhana Motherson] 데이터 추가: {quantity}")
            # else:
            #     print("숫자로 변환 불가")                        

# 기준: "Biesterfeld Polybass S.p.A."
def process_BiesterfeldPolybassSpA(text, filename, ws):
    """Biesterfeld Polybass S.p.A. 문서를 처리하는 함수"""
    print(f"📄 Processing file: {filename}")

    lines = text.replace(",", "").replace(".", "").split("\n")
    cleaned_list = [item.strip() for item in lines if item.strip()]
    written_date = ""
    written_month = ""
    pno = ""
    i = 0
    merge_lines = []
    while i < len(cleaned_list):
        if cleaned_list[i].strip().startswith("Currency") and i + 1 < len(cleaned_list):
            if cleaned_list[i+2].strip().startswith("FCA") and i + 1 < len(cleaned_list):
                merge_lines.append(f"{cleaned_list[i]} {cleaned_list[i+1]} Seob_Date")
            i+=2
        elif cleaned_list[i].strip().startswith("kg") and i + 1 < len(cleaned_list):
            merge_lines.append(f"{cleaned_list[i-4]} {cleaned_list[i]}")
            i+=1
        elif cleaned_list[i].strip().startswith("PO Number") and i + 1 < len(cleaned_list):
            merge_lines.append(f"{cleaned_list[i-1]} {cleaned_list[i]}")
            i+=1
        else:
            merge_lines.append(str(cleaned_list[i]))
            i += 1

    for line in merge_lines:
        extracted_texts = re.split(r'\s+', line)
        extracted_text = [item for item in extracted_texts if item.strip()]
        # print(f"@@@@@@@@@@@@@@@line {line}")
        # print(f"@@@@@@@@@@@@@@@extracted_text {extracted_text}")
        material = get_materialcode("5335630000") 

        if "Seob_Date" in line and len(extracted_text) >= 2:
            written_date = f"{extracted_text[1][:2]}-{extracted_text[1][2:4]}-{extracted_text[1][4:]}"
            written_month = f"{mon(extracted_text[1][2:4])}-{extracted_text[1][6:]}"
        
        if line.endswith("Number") and "PO" in line:
            if int(extracted_text[0]) > 0:
                pno = extracted_text[0]
        
        if line.endswith("kg") and len(extracted_text) >= 2:
            quantity = extracted_text[0]
            if quantity.isdigit() and int(quantity) > 0:
                qty = float(quantity)/1000

            first_two_lows("Biesterfeld Polybass S.p.A.", ws, material)
            ws.append([written_month, The_King, The_King, The_King, "On Stock", qty, written_date, pno, filename])    

# 기준: "Hagstrasse 1"
def process_Hagstrasse1(text, filename, ws):
    """Hagstrasse 1 문서를 처리하는 함수"""
    print(f"📄 Processing file: {filename}")

    lines = text.replace(",", "").replace(".", "").split("\n")
    cleaned_list = [item.strip() for item in lines if item.strip()]
    written_date = ""
    written_month = ""
    pno = ""
    i = 0
    merge_lines = []
    while i < len(cleaned_list):
        if "ABS LG ER400 M95007" in cleaned_list[i].strip():
            merge_lines.append(f"{cleaned_list[i]} {cleaned_list[i+2]} {cleaned_list[i+5]} {cleaned_list[i+6]}")
            i+=6
        elif "Bestelldatum" in cleaned_list[i].strip():
            merge_lines.append(f"{cleaned_list[i-1]} {cleaned_list[i]} {cleaned_list[i+1]}")
            i+=2
        else:
            merge_lines.append(str(cleaned_list[i]))
            i += 1
    
    for line in merge_lines:
        extracted_texts = re.split(r'\s+', line)
        extracted_text = [item for item in extracted_texts if item.strip()]
        material = get_materialcode("ABS LG ER400 M95007 schwarz")
        
        if line.startswith("ABS LG ER400"):
            if len(extracted_text) >= 5:
                quantity = extracted_text[4]
                if quantity.isdigit() and int(quantity) > 0:
                    qty = float(quantity)/1000
                    written_date = f"{extracted_text[6][:2]}-{extracted_text[6][2:4]}-{extracted_text[6][4:]}"
                    written_month = f"{mon(extracted_text[6][2:4])}-{extracted_text[6][4:]}"
                    ws.append([written_month, The_King, The_King, The_King, "On Stock", qty, written_date, "pno", filename])
                    first_two_lows("Hagstrasse 1", ws, material)
                    # print(f"[Samvardhana Motherson] 데이터 추가: {quantity}")
        if "Bestelldatum" in line:
            if len(extracted_text) >= 3:
                pno = extracted_text[0]
                first_row = 6
                last_row = ws.max_row
                column = 8
                for row in range(first_row, last_row+1):
                    cell = ws.cell(row=row, column=column)
                    cell.value = pno          
        
        
# 기준: "Maflow Plastics Poland"
def process_MaflowPlastic(text, filename, ws):
    """Maflow Plastic Poland 문서를 처리하는 함수"""
    print(f"📄 Processing file: {filename}")

    lines = text.replace(",", "").replace(".", "").split("\n")
    cleaned_list = [item.strip() for item in lines if item.strip()]
    i = 0
    merge_lines = []
    Weeks = ["Seob_First_Week"]
    Quantities = ["Seob_Quantity"]
    entries = []
    while i < len(cleaned_list):
        for The_Week in range(0,52):
            if f"W{The_Week}" in cleaned_list[i].strip():
                if "Supplied" in cleaned_list[i].strip():
                    First_Week = cleaned_list[i].strip()[-3:]
                    Weeks.append(f" {First_Week}")
                elif len(cleaned_list[i].strip()) == len(f"W{The_Week}"):
                    Weeks.append(f" W{The_Week}")
                elif len(cleaned_list[i].strip().split()) >= 2:
                    long_weeks = ""
                    long_weeks = cleaned_list[i].strip().split()
                    for each_week in long_weeks:
                        if f"W{The_Week}" == each_week:
                            Weeks.append(f" W{The_Week}")
        if "DOSTAWCA" in cleaned_list[i].strip() and i + 1 < len(cleaned_list):
            merge_lines.append(f"{cleaned_list[i]} {cleaned_list[i+1]} Seob_pno")
            i+=2
        elif "LG CHEMICAL LI 912 (ASA)" in cleaned_list[i].strip() and i + 1 < len(cleaned_list):
            for The_Quantity in range(i+2, len(cleaned_list)):
                Quantities.append(f" {cleaned_list[The_Quantity]}")
            i+=1
        else:
            merge_lines.append(str(cleaned_list[i]))
            i+=1

    cleaned_quantities = "".join(Quantities)
    cleaned_weeks = "".join(Weeks)

    merge_lines.append(cleaned_quantities)
    merge_lines.append(cleaned_weeks)
    
    for line in merge_lines:
        extracted_texts = re.split(r'\s+', line)
        extracted_text = [item for item in extracted_texts if item.strip()]
        material = get_materialcode("0075A00054100GR")
        # print(f"@@@@@@@@@@@@@@@line {line}")
        # print(f"@@@@@@@@@@@@@@@extracted_text {extracted_text}")

        if "Seob_pno" in line:
            if len(extracted_text) >= 2:
                pno = extracted_text[1]
                first_row = 6
                last_row = ws.max_row
                column = 8
                for row in range(first_row, last_row+1):
                    cell = ws.cell(row=row, column=column)
                    cell.value = pno 
        
        if "Seob_Quantity" in line:
            Seob_Quantity = extracted_text[1:]
        if "Seob_First_Week" in line:
            # 초기화
            New_Week = []
            New_Month = []

            # 주차 정보 리스트 추출 (extracted_text[1:] → ["W10", "W11", ...])
            Seob_Week = extracted_text[1:]

            for w_number in Seob_Week:
                # 주 번호만 숫자로 추출 (ex: "W10" → 10)
                week_num = int(w_number[1:])
                
                # 해당 주의 월요일 날짜 구하기
                monday = datetime.strptime(f"{datetime.now().year}-W{week_num}-1", "%G-W%V-%u")
                
                # 일-월-년 형식으로 변경
                date_str = monday.strftime("%d-%m-%Y")
                
                # 첫 열을 위해 열 표기 형식 변경
                month_str = f"{mon(date_str[3:5])}-{date_str[8:]}"

                # 리스트에 저장
                New_Week.append(date_str)
                New_Month.append(month_str)

            # 수량과 주차 수가 동일한 경우에만 entries에 추가
            if len(Seob_Quantity) == len(New_Week):
                for qty, written_date, written_month in zip(Seob_Quantity, New_Week, New_Month):
                    entries.append({"qty": qty, "written_date": written_date, "written_month": written_month, "pno": pno})

    for entry in entries:
        if int(entry["qty"]) > 0 and entry["written_date"] and entry["pno"]:
            ws.append([entry["written_month"],
                The_King,
                The_King,
                The_King,
                "On Stock",
                int(entry["qty"])/1000,
                entry["written_date"],
                entry["pno"],
                filename
            ])
    first_two_lows("Maflow Plastics Poland", ws, material)
            # print(f"✅ 추가됨: {entry['qty']}, {entry['written_date']}, {entry['pno']}")
        # else:
        #     print(f"❌ 정보 부족 → 건너뜀: {entry}")       

# 기준: "ITW Slovakia s.r.o."
def process_ITWSlovakiasro(text, filename, ws):
    """ITW Slovakia s.r.o. 문서를 처리하는 함수"""
    print(f"📄 Processing file: {filename}")

    lines = text.replace(",", "").replace(".", "").split("\n")
    cleaned_list = [item.strip() for item in lines if item.strip()]
    written_date = ""
    written_month = ""
    pno = ""
    i = 0
    merge_lines = []
    while i < len(cleaned_list):
        if cleaned_list[i].strip().endswith(" P") and i + 1 < len(cleaned_list):
            if cleaned_list[i+4].strip().endswith(" P") and i + 1 < len(cleaned_list):
                merge_lines.append(f"{cleaned_list[i]} {cleaned_list[i+1]} {cleaned_list[i+2]} {cleaned_list[i+3]} Seob4")
                i+=4
            elif cleaned_list[i+4].strip().endswith("Sincerly"):
                merge_lines.append(f"{cleaned_list[i]} {cleaned_list[i+1]} {cleaned_list[i+2]} {cleaned_list[i+3]} Seob4")
                i+=4
            else:
                merge_lines.append(f"{cleaned_list[i]} {cleaned_list[i+1]} {cleaned_list[i+2]} {cleaned_list[i+3]} {cleaned_list[i+4]} Seob5")
                i+=5
        elif cleaned_list[i].strip().startswith("Number:") and i + 1 < len(cleaned_list):
            merge_lines.append(f"{cleaned_list[i]} {cleaned_list[i+1]}")
            i+=2
        else:
            merge_lines.append(str(cleaned_list[i]))
            i += 1
    
    for line in merge_lines:
        extracted_texts = re.split(r'\s+', line)
        extracted_text = [item for item in extracted_texts if item.strip()]
        material = get_materialcode("None")

        if "Number:" in line:
            if len(extracted_text) >= 2:
                pno = (f"{extracted_text[1]} {extracted_text[2]}")
                first_row = 6
                last_row = ws.max_row
                column = 8
                for row in range(first_row, last_row+1):
                    cell = ws.cell(row=row, column=column)
                    cell.value = pno 
        
        if line.endswith("Seob4"):
            if len(extracted_text) >= 5:
                quantity = extracted_text[3]
                if quantity.isdigit() and int(quantity) > 0:
                    qty = float(quantity)/100000
                    if len(extracted_text[0]) == 5:
                        for_date = f"0{extracted_text[0]}"
                    else:
                        for_date = extracted_text[0]
                    written_date = f"{for_date[:2]}-{for_date[2:4]}-{for_date[4:]}"
                    written_month = f"{mon(for_date[2:4])}-{for_date[4:]}"
                    if written_date and pno:
                        # print(f"✅ 저장된 정보 기반 데이터 추가: {qty}, {written_date}, {pno}")
                        ws.append([written_month, The_King, The_King, The_King, "On Stock", qty, written_date, pno, filename]) 
                    # else:
                    #     print("❌ 날짜나 PNO 정보가 누락됨 → 건너뜀")

        if line.endswith("Seob5"):
            if len(extracted_text) >= 6:
                quantity = extracted_text[4]
                if quantity.isdigit() and int(quantity) > 0:
                    qty = float(quantity)/100000
                    if len(extracted_text[0]) == 5:
                        for_date = f"0{extracted_text[0]}"
                    else:
                        for_date = extracted_text[0]
                    written_date = f"{for_date[:2]}-{for_date[2:4]}-{for_date[4:]}"
                    written_month = f"{mon(for_date[2:4])}-{for_date[4:]}"
                    if written_date and pno:
                        # print(f"✅ 저장된 정보 기반 데이터 추가: {qty}, {written_date}, {pno}")
                        ws.append([written_month, The_King, The_King, The_King, "On Stock", qty, written_date, pno, filename]) 
        first_two_lows("ITW Slovakia s.r.o.", ws, material)
                    # else:
                    #     print("❌ 날짜나 PNO 정보가 누락됨 → 건너뜀")

# 기준: "Boryszew Kunststofftechnik"
def process_BoryszewKunststofftechnik(text, filename, ws):
    """Boryszew Kunststoff technik 문서를 처리하는 함수"""
    print(f"📄 Processing file: {filename}")

    lines = text.replace(",", "").replace(".", "").split("\n")
    cleaned_list = [item.strip() for item in lines if item.strip()]
    written_date = ""
    written_month = ""
    pno = ""
    i = 0
    merge_lines = []
    while i < len(cleaned_list):
        if cleaned_list[i].strip().endswith("Gesamtpreis") and i + 1 < len(cleaned_list):
            merge_lines.append(f"{cleaned_list[i]} {cleaned_list[i+3]} {cleaned_list[i+6]} Seob_Gesamtpreis")
            i+=6
        elif "Liefertermin:" in cleaned_list[i] and i + 1 < len(cleaned_list):
            merge_lines.append(f"{cleaned_list[i]} {cleaned_list[i+1]}")
            i+=2
        elif cleaned_list[i].strip().startswith("Hiermit bestellen") and i + 1 < len(cleaned_list):
            merge_lines.append(f"{cleaned_list[i-2]} {cleaned_list[i]} Seob_PO")
            i+=1
        else:
            merge_lines.append(str(cleaned_list[i]))
            i += 1
    
    for line in merge_lines:
        extracted_texts = re.split(r'\s+', line)
        extracted_text = [item for item in extracted_texts if item.strip()]
        material = get_materialcode("ABS XR410 naturverpackt in Octabin")
        # print(f"@@@@@@@@@@@@@@@line {line}")
        # print(f"@@@@@@@@@@@@@@@extracted_text {extracted_text}")

        if "Seob_PO" in line:
            if len(extracted_text) >= 2:
                pno = (f"{extracted_text[0]}")
                first_row = 6
                last_row = ws.max_row
                column = 8
                for row in range(first_row, last_row+1):
                    cell = ws.cell(row=row, column=column)
                    cell.value = pno 

        if "Liefertermin:" in line:
            for_date = extracted_text[1]
            written_date = f"{for_date[:2]}-{for_date[2:4]}-{for_date[4:]}"
            written_month = f"{mon(for_date[2:4])}-{for_date[6:]}"
            for row in range(6, ws.max_row + 1):
                    cell = ws.cell(row=row, column=7)
                    cell2 = ws.cell(row=row, column=1)
                    cell.value = written_date 
                    cell2.value = written_month

       
        if "Seob_Gesamtpreis" in line:
            if len(extracted_text) >= 4:
                quantity = extracted_text[2]
                if quantity.isdigit() and int(quantity) > 0:
                    qty = float(quantity)/1000
                    ws.append([written_month, The_King, The_King, The_King, "On Stock", qty, written_date, pno, filename])
                    first_two_lows("Boryszew Kunststofftechnik", ws, material)
                    # print(f"[ITWSlovakia] 데이터 추가: {quantity}")         

# 기준: "PCZ-571 01 MORAVSKA TREBOVA"
def process_ProXMORAVSKATREBOVA(text, filename, ws):
    """CZ-571 01 MORAVSKA TREBOVA 문서를 처리하는 함수"""
    print(f"📄 Processing file: {filename}")

    lines = text.replace(",", "").replace(".", "").split("\n")
    cleaned_list = [item.strip() for item in lines if item.strip()]
    merge_lines = []

    # 임시 저장용 변수들
    written_date = ""
    written_month = ""
    pno = ""
    i = 0
    while i < len(cleaned_list):
        if "Bestellnummer / Datum" in cleaned_list[i] and i + 1 < len(cleaned_list):
            merge_lines.append(f"{cleaned_list[i]} {cleaned_list[i+1]} Seob_Bestellnummer")
            i += 2
        elif "Lieferdatum" in cleaned_list[i] and i + 1 < len(cleaned_list):
            merge_lines.append(f"{cleaned_list[i]} {cleaned_list[i+1]} Seob_Date")
            i += 2
        elif "LG ASA LI941" in cleaned_list[i] and i + 1 < len(cleaned_list):
            merge_lines.append(f"{cleaned_list[i]} {cleaned_list[i+1]} Seob_qty")
            i += 2
        else:
            merge_lines.append(cleaned_list[i])
            i += 1

    for line in merge_lines:
        extracted_texts = re.split(r'\s+', line)
        extracted_text = [item for item in extracted_texts if item.strip()]
        if "LG ASA LI941-V - 94841 (VW9B9) (SILO)" in line:
            material = get_materialcode("LG ASA LI941-V - 94841 (VW9B9) (SILO)")
            C_Name = "Pro-X MORAVSKA TREBOVA"
        elif "LG ASA LI941-F - 94841 (VW9B9) AEB" in line:
            material = get_materialcode("LG ASA LI941-F - 94841 (VW9B9) AEB")
            C_Name = "Pro-X FEUCHTWANGEN"
        elif "LG ASA LI941 - V94841 (VW9B9) BigBag" in line:
            material = get_materialcode("LG ASA LI941 - V94841 (VW9B9) BigBag")
            C_Name = "Pro-X  Eckerle Spritz"
        # print(f"@@@@@@@@@@@@@@@line {line}")
        # print(f"@@@@@@@@@@@@@@@extracted_text {extracted_text}")

        if "Seob_Bestellnummer" in line and len(extracted_text) >= 8:
            pno = extracted_text[5]
            # print(f"📌 날짜/PNO 저장: {written_date}, {pno}")

        if "Seob_Date" in line and len(extracted_text) >= 2:
            for_date = extracted_text[1]
            written_date = f"{for_date[:2]}-{for_date[2:4]}-{for_date[4:]}"
            written_month = f"{mon(for_date[2:4])}-{for_date[6:]}"

        if "Seob_qty" in line and len(extracted_text) >= 9:
            quantity = extracted_text[7]
            if quantity.isdigit() and int(quantity) > 0:
                qty = float(quantity) / 1000000
                if written_date and pno:
                    # print(f"✅ 저장된 정보 기반 데이터 추가: {qty}, {written_date}, {pno}")
                    ws.append([written_month, The_King, The_King, The_King, "On Stock", qty, written_date, pno, filename])
                    first_two_lows(C_Name, ws, material)
                # else:
                    # print("❌ 날짜나 PNO 정보가 누락됨 → 건너뜀")

# 기준: "Finke Anwendungstechnik GmbH"
def process_FinkeAnwendungstechnik(text, filename, ws):
    """Finke Anwendungstechnik GmbH 문서를 처리하는 함수"""
    print(f"📄 Processing file: {filename}")

    lines = text.replace(",", "").replace(".", "").split("\n")
    cleaned_list = [item.strip() for item in lines if item.strip()]
    i = 0
    merge_lines = []
    written_date = ""
    written_month = ""
    pno = ""
    entries = []  # 여기서 모든 qty와 그 시점의 context(pno, date 등)를 저장

    while i < len(cleaned_list):
        line = cleaned_list[i]
        if "Bestellung Nr:" in line and i + 1 < len(cleaned_list):
            merge_lines.append(f"{cleaned_list[i]} {cleaned_list[i+1]} Seob_BestellungNr")
            i += 2
        elif "Liefertermin:" in line and i + 1 < len(cleaned_list):
            merge_lines.append(f"{cleaned_list[i]} Seb_Date")
            i += 1
        elif "Gesamt Netto" in line and i + 1 < len(cleaned_list):
            end = i - 5
            for kg in range(i, end, -1):
                parts = cleaned_list[kg].split()
                if len(parts) >= 2 and parts[1].upper() == "KG":
                    merge_lines.append(f"{cleaned_list[kg]} Seob_qty")
            i += 1
        else:
            merge_lines.append(line)
            i += 1

    for line in merge_lines:
        extracted_texts = re.split(r'\s+', line)
        extracted_text = [item for item in extracted_texts if item.strip()]
        # print(f"@@@@@@@@@@@@@@@line {line}")
        # print(f"@@@@@@@@@@@@@@@extracted_text {extracted_text}")
        if "High gloss ASA: LI941F" in line:
            material = get_materialcode("High gloss ASA: LI941F Piano Black (F94484)")
            # first_two_lows("Finke Anwendungstechnik GmbH", ws, material)
        elif "ABS XR 401 BK 9001" in line:
            material = get_materialcode("ABS XR 401 BK 9001")

        if "Seob_BestellungNr" in line:
            if len(extracted_text) >= 5:
                pno = extracted_text[2]
        elif "Seob_qty" in line:
            if len(extracted_text) >= 2:
                quantity = extracted_text[0]
                if quantity.isdigit() and int(quantity) > 0:
                    qty = float(quantity) / 100000
        elif "Seb_Date" in line:
            if len(extracted_text) >= 2:
                for_date = extracted_text[1]
                written_date = f"{for_date[:2]}-{for_date[2:4]}-{for_date[4:]}"
                written_month = f"{mon(for_date[2:4])}-{for_date[6:]}"
                entries.append({"qty": qty, "written_date": written_date, "written_month": written_month, "pno": pno})

    for entry in entries:
        if entry["written_date"] and entry["pno"]:
            ws.append([entry["written_month"],
                The_King,
                The_King,
                The_King,
                "On Stock",
                entry["qty"],
                entry["written_date"],
                entry["pno"],
                filename
            ])
    first_two_lows("Finke Anwendungstechnik GmbH", ws, material)
            # print(f"✅ 추가됨: {entry['qty']}, {entry['written_date']}, {entry['pno']}")
        # else:
        #     print(f"❌ 정보 부족 → 건너뜀: {entry}")

# 기준: "Formzeug GmbH"
def process_FormzeugGmbH(text, filename, ws):
    """Formzeug GmbH 문서를 처리하는 함수"""
    print(f"📄 Processing file: {filename}")

    lines = text.replace(",", "").replace(".", "").split("\n")
    cleaned_list = [item.strip() for item in lines if item.strip()]
    written_date = ""
    written_month = ""
    pno = ""
    i = 0
    merge_lines = []
    while i < len(cleaned_list):
        if "Bestellung " in cleaned_list[i].strip() and i + 1 < len(cleaned_list):
            merge_lines.append(f"{cleaned_list[i]} {cleaned_list[i+1]} Seob_BestellungNr")
            i+=2
        elif "ASA LI941-F94484" in cleaned_list[i].strip() and i + 1 < len(cleaned_list):
            merge_lines.append(f"{cleaned_list[i-2]} {cleaned_list[i]} Seob_qty")
            i+=1
        elif "Lieferanschrift" in cleaned_list[i].strip() and i + 1 < len(cleaned_list):
            merge_lines.append(f"{cleaned_list[i-2]} Seob_date")
            i+=1
        else:
            merge_lines.append(str(cleaned_list[i]))
            i += 1
    
    for line in merge_lines:
        extracted_texts = re.split(r'\s+', line)
        extracted_text = [item for item in extracted_texts if item.strip()]
        material = get_materialcode("High gloss ASA LI941-F94484 (Piano Black)")
        # print(f"@@@@@@@@@@@@@@@line {line}")
        # print(f"@@@@@@@@@@@@@@@extracted_text {extracted_text}")

        if "Seob_BestellungNr" in line:
            pno = extracted_text[1]
            for row in range(6, ws.max_row + 1):
                    cell_pno = ws.cell(row=row, column=8)
                    cell_pno.value = pno

        
       
        if "Seob_qty" in line:
            if len(extracted_text) >= 5:
                quantity = extracted_text[0]
                if quantity.isdigit() and int(quantity) > 0:
                    qty = float(quantity)/1000

        if "Seob_date" in line:
            for_date = extracted_text[0]
            written_date = f"{for_date[:2]}-{for_date[2:4]}-{for_date[4:]}"
            written_month = f"{mon(for_date[2:4])}-{for_date[6:]}"
                    # print(f"✅ 저장된 정보 기반 데이터 추가: {qty}, {written_date}, {pno}")
            ws.append([written_month, The_King, The_King, The_King, "On Stock", qty, written_date, pno, filename]) 
            first_two_lows("Formzeug GmbH", ws, material)
                    # else:
                    #     print("❌ 날짜나 PNO 정보가 누락됨 → 건너뜀")

# 기준: "ABC Technologies Karl Etzel GmbH"
def process_ABCTechnologiesKEGmbH(text, filename, ws):
    """ABC Technologies Karl Etzel GmbH 문서를 처리하는 함수"""
    print(f"📄 Processing file: {filename}")

    lines = text.replace(",", "").replace(".", "").split("\n")
    cleaned_list = [item.strip() for item in lines if item.strip()]
    i = 0
    merge_lines = []
    written_date = ""
    written_month = ""
    pno = ""
    entries = []  # 여기서 모든 qty와 그 시점의 context(pno, date 등)를 저장

    while i < len(cleaned_list):
        line = cleaned_list[i]
        if "Belegnummer:" in line and i + 1 < len(cleaned_list):
            merge_lines.append(f"{cleaned_list[i-5]} {cleaned_list[i]} Seob_BestellungNr")
            i += 1
        elif "ABS062" in line and i + 1 < len(cleaned_list):
            merge_lines.append(f"{cleaned_list[i]} {cleaned_list[i+1]} Seob_qty")
            i += 1
        else:
            merge_lines.append(line)
            i += 1

    for line in merge_lines:
        extracted_texts = re.split(r'\s+', line)
        extracted_text = [item for item in extracted_texts if item.strip()]
        material = get_materialcode("ABS ER400 M95007 schwarz")
        # print(f"@@@@@@@@@@@@@@@line {line}")
        # print(f"@@@@@@@@@@@@@@@extracted_text {extracted_text}")

        if "Seob_BestellungNr" in line:
            if len(extracted_text) >= 2:
                pno = extracted_text[0]

        if "Seob_qty" in line:
            if len(extracted_text) >= 3:
                quantity = extracted_text[1]
                if quantity.isdigit() and int(quantity) > 0:
                    qty = float(quantity) / 100000
                    
        if "Liefertermin:" in line and len(extracted_text) == 2:
            for_date = extracted_text[1]
            written_date = f"{for_date[:2]}-{for_date[2:4]}-{for_date[4:]}"
            written_month = f"{mon(for_date[2:4])}-{for_date[6:]}"
            if pno and written_date:
                entries.append({"qty": qty, "written_date": written_date, "written_month": written_month, "pno": pno})
                # print(f"✅ qty 추가됨: {qty}, {written_date}, {pno}")
            # else:
            #     pending_qty = qty
            #     print(f"⏸ 날짜 없음, pending에 저장: {pending_qty}")

    for entry in entries:
        if entry["written_date"] and entry["pno"]:
            ws.append([entry["written_month"],
                The_King,
                The_King,
                The_King,
                "On Stock",
                entry["qty"],
                entry["written_date"],
                entry["pno"],
                filename
            ])

    first_two_lows("ABC Technologies Karl Etzel GmbH", ws, material)
            # print(f"✅ 추가됨: {entry['qty']}, {entry['written_date']}, {entry['pno']}")
        # else:
        #     print(f"❌ 정보 부족 → 건너뜀: {entry}")

# 기준: "Plant Oldenburg"
def process_PlantOldenburg(text, filename, ws):
    """Plant Oldenburg 문서를 처리하는 함수"""
    print(f"📄 Processing file: {filename}")

    lines = text.replace(",", "").replace(".", "").split("\n")
    cleaned_list = [item.strip() for item in lines if item.strip()]
    i = 0
    merge_lines = []
    written_date = ""
    written_month = ""
    pno = ""

    while i < len(cleaned_list):
        line = cleaned_list[i]
        temp_text = ""
        if "Lieferplannummer/Datum" in line and i + 1 < len(cleaned_list):
            merge_lines.append(f"{cleaned_list[i]} {cleaned_list[i+1]} Seob_BestellungNr")
            i += 2
        elif cleaned_list[i][0] == "T" and cleaned_list[i][6:8] == "20":
            merge_lines.append(f"{cleaned_list[i]} {cleaned_list[i+1]} Seob_qty")
            i += 2
        # 뻘짓이었으나 로직이 아까워서 살림. 아래것
        elif "KG" in line and i + 1 < len(cleaned_list):
            temp_text+= cleaned_list[i-1] + " " + cleaned_list[i+1]
            D_Date = temp_text.split()
            if "202" in D_Date[2].strip() and int(D_Date[0]) > 0:
                merge_lines.append(f"{temp_text}")
            i += 2
        else:
            merge_lines.append(line)
            i += 1

    for line in merge_lines:
        extracted_texts = re.split(r'\s+', line)
        extracted_text = [item for item in extracted_texts if item.strip()]
        material = get_materialcode("30022028 LG LI941-F 94484 PIANO BLACK")
        # print(f"@@@@@@@@@@@@@@@line {line}")
        # print(f"@@@@@@@@@@@@@@@extracted_text {extracted_text}")

        if "Seob_qty" in line:
            if len(extracted_text) >= 3:
                quantity = extracted_text[2]
                for_date = extracted_text[1]
                written_date = f"{for_date[:2]}-{for_date[2:4]}-{for_date[4:]}"
                written_month = f"{mon(for_date[2:4])}-{for_date[6:]}"
                if quantity.isdigit() and int(quantity) > 0:
                    qty = float(quantity) / 1000000
                    ws.append([written_month, The_King, The_King, The_King, "On Stock", qty, written_date, pno, filename])

        if "Seob_BestellungNr" in line:
            if len(extracted_text) >= 2:
                pno = extracted_text[1]
                first_row = 6
                last_row = ws.max_row
                column = 8
                for row in range(first_row, last_row+1):
                    cell = ws.cell(row=row, column=column)
                    cell.value = pno  
    first_two_lows("Plant Oldenburg", ws, material)

# 기준: "c/o Linden GmbH"
def process_coLindenGmbH(text, filename, ws):
    """c/o Linden GmbH 문서를 처리하는 함수"""
    print(f"📄 Processing file: {filename}")

    lines = text.replace(",", "").replace(".", "").split("\n")
    cleaned_list = [item.strip() for item in lines if item.strip()]
    i = 0
    merge_lines = []
    written_date = ""
    written_month = ""
    pno = ""

    while i < len(cleaned_list):
        line = cleaned_list[i]
        if "Lieferplannummer/Datum" in line and i + 1 < len(cleaned_list):
            merge_lines.append(f"{cleaned_list[i]} {cleaned_list[i+1]} Seob_BestellungNr")
            i += 2
        elif "number/date" in line  and i + 1 < len(cleaned_list):
            merge_lines.append(f"Seob_BestellungNr {cleaned_list[i+1]} Seob_BestellungNr")
            i+=2
        elif "KG" in line and i + 1 < len(cleaned_list):
            temp_first = ""
            if cleaned_list[i-2][0] == "T":
                merge_lines.append(f"{cleaned_list[i-2]} {cleaned_list[i-1]} {cleaned_list[i]} Seob_qty")
            elif cleaned_list[i-2][0] == "M":
                Monday = first_monday(cleaned_list[i-2][1:4], cleaned_list[i-2][4:])
                temp_first = cleaned_list[i-2][0] + " 0" + str(Monday) + cleaned_list[i-2][1:].replace(" ","")
                merge_lines.append(f"{temp_first} {cleaned_list[i-1]} {cleaned_list[i]} Seob_qty")
            i += 1
        else:
            merge_lines.append(line)
            i += 1

    for line in merge_lines:
        extracted_texts = re.split(r'\s+', line)
        extracted_text = [item for item in extracted_texts if item.strip()]
        if "30021896 LG LI941V 94841" in line:
            material = get_materialcode("30021896 LG LI941V 94841")
        elif "30022062 ABS XR 410 NATUR" in line:
            material = get_materialcode("30022062 ABS XR 410 NATUR")
        # print(f"@@@@@@@@@@@@@@@line {line}")
        # print(f"@@@@@@@@@@@@@@@extracted_text {extracted_text}")

        if "Seob_qty" in line:
            if len(extracted_text) >= 5:
                quantity = extracted_text[3]
                for_date = extracted_text[1]
                written_date = f"{for_date[:2]}-{for_date[2:4]}-{for_date[4:]}"
                written_month = f"{mon(for_date[2:4])}-{for_date[6:]}"
                if quantity.isdigit() and int(quantity) > 0:
                    qty = float(quantity) / 1000000
                    ws.append([written_month, The_King, The_King, The_King, "On Stock", qty, written_date, pno, filename])

        if "Seob_BestellungNr" in line:
            if len(extracted_text) >= 3:
                pno = extracted_text[1]
                first_row = 6
                last_row = ws.max_row
                column = 8
                for row in range(first_row, last_row+1):
                    cell = ws.cell(row=row, column=column)
                    cell.value = pno  
    first_two_lows("C/O Linden GmbH", ws, material)

# 기준: "Plant Neustadt"
def process_PlantNeustadt(text, filename, ws):
    """Plant Neustadt 문서를 처리하는 함수"""
    print(f"📄 Processing file: {filename}")

    lines = text.replace(",", "").replace(".", "").split("\n")
    cleaned_list = [item.strip() for item in lines if item.strip()]
    i = 0
    merge_lines = []
    written_date = ""
    written_month = ""
    pno = ""

    while i < len(cleaned_list):
        line = cleaned_list[i]
        if "Purchase Order:" in line and i + 1 < len(cleaned_list):
            merge_lines.append(f"{cleaned_list[i]} Seob_BestellungNr")
            i += 1
        elif "ABS 950kg XR410 9529" in line and i + 1 < len(cleaned_list):
            if "kg" in cleaned_list[i+1] and "Delivery date:" in cleaned_list[i+5]:
                merge_lines.append(f"{cleaned_list[i]} {cleaned_list[i+1]} {cleaned_list[i+5]} Seob_qty")
            i += 1
        else:
            merge_lines.append(line)
            i += 1

    for line in merge_lines:
        extracted_texts = re.split(r'\s+', line)
        extracted_text = [item for item in extracted_texts if item.strip()]
        material = get_materialcode("ABS 950kg XR410 9529")
        # print(f"@@@@@@@@@@@@@@@line {line}")
        # print(f"@@@@@@@@@@@@@@@extracted_text {extracted_text}")

        if "Seob_qty" in line:
            if len(extracted_text) >= 9:
                quantity = extracted_text[4]
                for_date = extracted_text[8]
                written_date = f"{for_date[9:]}-{rev_mon(for_date[5:8])}-{for_date[:4]}"
                written_month = f"{for_date[5:8]}-{for_date[2:4]}"
                if quantity.isdigit() and int(quantity) > 0:
                    qty = float(quantity) / 100000
                    ws.append([written_month, The_King, The_King, The_King, "On Stock", qty, written_date, pno, filename])

        if "Seob_BestellungNr" in line:
            if len(extracted_text) >= 3:
                pno = extracted_text[2]
                first_row = 6
                last_row = ws.max_row
                column = 8
                for row in range(first_row, last_row+1):
                    cell = ws.cell(row=row, column=column)
                    cell.value = pno  
    first_two_lows("Plant Neustadt", ws, material)

# Material Code = "ASA LI941 F94484 (LG)" 경우에는 Iberica SLU 와 중복 값
# 기준: "SMP Automotive Technology Ibérica SLU"
def process_SMPAutomotiveTechnologyIbéricaSLU(text, filename, ws):
    """SMP Automotive Technology Ibérica SLU 문서를 처리하는 함수"""
    print(f"📄 Processing file: {filename}")

    lines = text.replace(",", "").replace(".", "").split("\n")
    cleaned_list = [item.strip() for item in lines if item.strip()]
    i = 0
    merge_lines = []
    written_date = ""
    written_month = ""
    pno = ""

    while i < len(cleaned_list):
        line = cleaned_list[i]
        if "Purchase Order:" in line and i + 1 < len(cleaned_list):
            merge_lines.append(f"{cleaned_list[i]} Seob_BestellungNr")
            i += 1
        elif "ASA LI941 F94484 (LG)" in line and i + 1 < len(cleaned_list):
            if "kg" in cleaned_list[i+1].lower() and "Delivery date:" in cleaned_list[i+5]:
                merge_lines.append(f"{cleaned_list[i]} {cleaned_list[i+1]} {cleaned_list[i+5]} Seob_qty")
            i += 2
        else:
            merge_lines.append(line)
            i += 1

    for line in merge_lines:
        extracted_texts = re.split(r'\s+', line)
        extracted_text = [item for item in extracted_texts if item.strip()]
        material = get_materialcode("ASA LI941 F94484 (LG)")
        # print(f"@@@@@@@@@@@@@@@line {line}")
        # print(f"@@@@@@@@@@@@@@@extracted_text {extracted_text}")

        if "Seob_qty" in line:
            if len(extracted_text) >= 9:
                quantity = extracted_text[4]
                for_date = extracted_text[8]
                written_date = f"{for_date[9:]}-{rev_mon(for_date[5:8])}-{for_date[:4]}"
                written_month = f"{for_date[5:8]}-{for_date[2:4]}"
                if quantity.isdigit() and int(quantity) > 0:
                    qty = float(quantity) / 100000
                    ws.append([written_month, The_King, The_King, The_King, "On Stock", qty, written_date, pno, filename])

        if "Seob_BestellungNr" in line:
            if len(extracted_text) >= 3:
                pno = extracted_text[2]
                first_row = 6
                last_row = ws.max_row
                column = 8
                for row in range(first_row, last_row+1):
                    cell = ws.cell(row=row, column=column)
                    cell.value = pno  
    first_two_lows("SMP Automotive Technology Ibérica SLU", ws, material)              

# 기준: "SMR Automotive Mirror Technology"
def process_SMRAutomotiveMirrorTechnology(text, filename, ws):
    """SMR Automotive Mirror Technology 문서를 처리하는 함수"""
    print(f"📄 Processing file: {filename}")

    lines = text.replace(",", "").replace(".", "").split("\n")
    cleaned_list = [item.strip() for item in lines if item.strip()]
    i = 0
    merge_lines = []
    written_date = ""
    written_month = ""
    pno = ""

    while i < len(cleaned_list):
        line = cleaned_list[i]
        if "number/date" in line and i + 1 < len(cleaned_list):
            merge_lines.append(f"{cleaned_list[i]} {cleaned_list[i+1]} Seob_BestellungNr")
            i += 2
        elif "T" in cleaned_list[i][0] and i + 1 < len(cleaned_list):
            print(cleaned_list[i][6:8])
            if "kg" in cleaned_list[i+2].lower() and cleaned_list[i][6:8] == "20":
                merge_lines.append(f"{cleaned_list[i]} {cleaned_list[i+1]} {cleaned_list[i+2]} Seob_qty")
            i += 3
        else:
            merge_lines.append(line)
            i += 1

    for line in merge_lines:
        extracted_texts = re.split(r'\s+', line)
        extracted_text = [item for item in extracted_texts if item.strip()]
        if "ASA LI941V" in line:
            material = get_materialcode("ASA LI941V")
        elif "ASA LI941F-94841" in line:
            material = get_materialcode("ASA LI941F-94841")
        elif "LG LI941 F 94484 PIANO BLACK" in line:
            material = get_materialcode("LG LI941 F 94484 PIANO BLACK")

        print(f"@@@@@@@@@@@@@@@line {line}")
        print(f"@@@@@@@@@@@@@@@extracted_text {extracted_text}")

        if "Seob_qty" in line:
            if len(extracted_text) >= 5:
                quantity = extracted_text[2]
                for_date = extracted_text[1]
                written_date = f"{for_date[:2]}-{for_date[2:4]}-{for_date[4:]}"
                written_month = f"{mon(for_date[2:4])}-{for_date[6:]}"
                if quantity.isdigit() and int(quantity) > 0:
                    qty = float(quantity) / 1000000
                    ws.append([written_month, The_King, The_King, The_King, "On Stock", qty, written_date, pno, filename])

        if "Seob_BestellungNr" in line:
            if len(extracted_text) >= 6:
                pno = extracted_text[3]
                first_row = 6
                last_row = ws.max_row
                column = 8
                for row in range(first_row, last_row+1):
                    cell = ws.cell(row=row, column=column)
                    cell.value = pno  
    first_two_lows("SMR Hungary Bt", ws, material)

# 기준: "Plant Schierling"
def process_PlantSchierling(text, filename, ws):
    """Plant Schierling 문서를 처리하는 함수"""
    print(f"📄 Processing file: {filename}")

    lines = text.replace(",", "").replace(".", "").split("\n")
    cleaned_list = [item.strip() for item in lines if item.strip()]
    i = 0
    merge_lines = []
    written_date = ""
    written_month = ""
    pno = ""

    while i < len(cleaned_list):
        line = cleaned_list[i]
        if "Lieferplannummer/Datum" in line and i + 1 < len(cleaned_list):
            merge_lines.append(f"{cleaned_list[i]} {cleaned_list[i+1]} Seob_BestellungNr")
            i += 2
        elif cleaned_list[i][0] == "T" and cleaned_list[i][6:8] == "20":
            merge_lines.append(f"{cleaned_list[i]} {cleaned_list[i+1]} Seob_qty")
            i += 2
        else:
            merge_lines.append(line)
            i += 1

    for line in merge_lines:
        extracted_texts = re.split(r'\s+', line)
        extracted_text = [item for item in extracted_texts if item.strip()]
        material = get_materialcode("ABS XR 410 NATUR")
        # print(f"@@@@@@@@@@@@@@@line {line}")
        # print(f"@@@@@@@@@@@@@@@extracted_text {extracted_text}")

        if "Seob_qty" in line:
            if len(extracted_text) >= 3:
                quantity = extracted_text[2]
                for_date = extracted_text[1]
                written_date = f"{for_date[:2]}-{for_date[2:4]}-{for_date[4:]}"
                written_month = f"{mon(for_date[2:4])}-{for_date[6:]}"
                if quantity.isdigit() and int(quantity) > 0:
                    qty = float(quantity) / 1000000
                    ws.append([written_month, The_King, The_King, The_King, "On Stock", qty, written_date, pno, filename])

        if "Seob_BestellungNr" in line:
            if len(extracted_text) >= 2:
                pno = extracted_text[1]
                first_row = 6
                last_row = ws.max_row
                column = 8
                for row in range(first_row, last_row+1):
                    cell = ws.cell(row=row, column=column)
                    cell.value = pno  
    first_two_lows("Plant Schierling", ws, material)

# 기준: "Uwe Etzel GmbH"
def process_UweEtzelGmbH(text, filename, ws):
    """Uwe Etzel GmbH 문서를 처리하는 함수"""
    print(f"📄 Processing file: {filename}")

    lines = text.replace(",", "").replace(".", "").split("\n")
    cleaned_list = [item.strip() for item in lines if item.strip()]
    i = 0
    merge_lines = []
    written_date = ""
    written_month = ""
    pno = ""

    while i < len(cleaned_list):
        line = cleaned_list[i]
        if "BESTELLUNG" in line and "Seite:" in cleaned_list[i+2] and i + 1 < len(cleaned_list):
            merge_lines.append(f"{cleaned_list[i]} {cleaned_list[i+1]} Seob_BestellungNr")
            i += 2
        elif cleaned_list[i].startswith("ABS032") and "kg" in cleaned_list[i+1].strip():
            merge_lines.append(f"{cleaned_list[i]} {cleaned_list[i+1]} Seob_qty")
            i += 2
        elif "Liefertermin:" in cleaned_list[i] and i + 1 < len(cleaned_list):
            merge_lines.append(f"{cleaned_list[i]} {cleaned_list[i+1]} Seob_Date")
            i += 2
        else:
            merge_lines.append(line)
            i += 1

    for line in merge_lines:
        extracted_texts = re.split(r'\s+', line)
        extracted_text = [item for item in extracted_texts if item.strip()]
        material = get_materialcode("ABS ER400-M95007")
        # print(f"@@@@@@@@@@@@@@@line {line}")
        # print(f"@@@@@@@@@@@@@@@extracted_text {extracted_text}")

        if "Seob_qty" in line:
            if len(extracted_text) >= 3:
                quantity = extracted_text[1]
                if quantity.isdigit() and int(quantity) > 0:
                    qty = float(quantity) / 100000           

        if "Seob_Date" in line:
            if len(extracted_text) >= 3:
                for_date = extracted_text[1]
                written_date = f"{for_date[:2]}-{for_date[2:4]}-{for_date[4:]}"
                written_month = f"{mon(for_date[2:4])}-{for_date[6:]}"
                ws.append([written_month, The_King, The_King, The_King, "On Stock", qty, written_date, pno, filename])

        if "Seob_BestellungNr" in line:
            if len(extracted_text) >= 2:
                pno = extracted_text[1]
                first_row = 6
                last_row = ws.max_row
                column = 8
                for row in range(first_row, last_row+1):
                    cell = ws.cell(row=row, column=column)
                    cell.value = pno  
    first_two_lows("Uwe Etzel GmbH", ws, material)

# 기준: "SLM Kunststofftechnik GmbH"
def process_SLMKunststofftechnikGmbH(text, filename, ws):
    """SLM Kunststofftechnik GmbH 문서를 처리하는 함수"""
    print(f"📄 Processing file: {filename}")

    lines = text.replace(",", "").replace(".", "").split("\n")
    cleaned_list = [item.strip() for item in lines if item.strip()]
    i = 0
    merge_lines = []
    written_date = ""
    written_month = ""
    pno = ""

    while i < len(cleaned_list):
        line = cleaned_list[i]
        if "Bestellnummer" in line and "Datum" in cleaned_list[i+2] and i + 1 < len(cleaned_list):
            merge_lines.append(f"{cleaned_list[i]} {cleaned_list[i+1]} Seob_BestellungNr")
            i += 2
        elif "101505" in cleaned_list[i].strip() and "kg" in cleaned_list[i+2].lower():
            merge_lines.append(f"{cleaned_list[i]} {cleaned_list[i+1]} Seob_qty")
            i += 2
        elif "Lieferdatum" in cleaned_list[i] and "Sehr" in cleaned_list[i+2] and i + 1 < len(cleaned_list):
            merge_lines.append(f"{cleaned_list[i]} {cleaned_list[i+1]} Seob_Date")
            i += 2
        else:
            merge_lines.append(line)
            i += 1

    for line in merge_lines:
        extracted_texts = re.split(r'\s+', line)
        extracted_text = [item for item in extracted_texts if item.strip()]
        material = get_materialcode("ASALI941-F94841 (9B9)")
        print(f"@@@@@@@@@@@@@@@line {line}")
        print(f"@@@@@@@@@@@@@@@extracted_text {extracted_text}")                   

        if "Seob_Date" in line:
            if len(extracted_text) >= 2:
                for_date = extracted_text[1]
                written_date = f"{for_date[:2]}-{for_date[2:4]}-{for_date[4:]}"
                written_month = f"{mon(for_date[2:4])}-{for_date[6:]}"

        if "Seob_qty" in line:
            if len(extracted_text) >= 3:
                quantity = extracted_text[2]
                if quantity.isdigit() and int(quantity) > 0:
                    qty = float(quantity) / 100000
                    ws.append([written_month, The_King, The_King, The_King, "On Stock", qty, written_date, pno, filename])

        if "Seob_BestellungNr" in line:
            if len(extracted_text) >= 2:
                pno = extracted_text[1]
                first_row = 6
                last_row = ws.max_row
                column = 8
                for row in range(first_row, last_row+1):
                    cell = ws.cell(row=row, column=column)
                    cell.value = pno  
    first_two_lows("SLM Kunststofftechnik GmbH", ws, material)




#########################################################
################ 회사별 함수 분류 끝 #####################
#########################################################


def extract_info(folder_path, output_excel):
    """폴더 내 모든 PDF를 읽고 키워드별로 처리"""
    extracted_data = []
    No_Key_Word = []

    # 기존 Excel 파일이 있으면 로드, 없으면 새 파일 생성
    if os.path.exists(output_excel):
        wb = openpyxl.load_workbook(output_excel)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # 기본 생성되는 'Sheet' 삭제

    for filename in os.listdir(folder_path):
        if filename.lower().endswith(".pdf"):
            file_path = os.path.join(folder_path, filename)

            try:
                doc = fitz.open(file_path)
                full_text = ""

                for page in doc:
                    text = page.get_text("text")
                    full_text+=text
                    # print(f"📄 {filename} - 페이지 텍스트 읽음")
                    # print(f"@@@@ Text 임: {text}")

                    if "SMP Ibérica S.L.U." in text and text != full_text:
                        ws = get_or_create_sheet(wb, "SMP IBERICA S.L.U. PALENCIA")
                        process_smp_iberica(full_text, filename, ws)

                    elif "Samvardhana Motherson Peguform" in text:
                        ws = get_or_create_sheet(wb, "Samvardhana Motherson Peguform")
                        process_samvardhanaPeguform(text, filename, ws)

                    elif "Samvardhana Motherson Innovative" in text:
                        ws = get_or_create_sheet(wb, "Samvardhana Motherson Innovative")
                        process_samvardhanaInnovative(text, filename, ws)

                    elif "Biesterfeld Polybass S.p.A." in text and "5335630000" in text:
                        ws = get_or_create_sheet(wb, "Biesterfeld Polybass S.p.A.")
                        process_BiesterfeldPolybassSpA(text, filename, ws)

                    elif "PLASTICOS ABC SPAIN" in text:
                        ws = get_or_create_sheet(wb, "PLASTICOS ABC SPAIN")
                        process_PLASTICOS(text, filename, ws)

                    elif "Hagstrasse 1" in text:
                        ws = get_or_create_sheet(wb, "Hagstrasse 1")
                        process_Hagstrasse1(text, filename, ws)
                        
                    elif "Maflow Plastics Poland" in text:
                        ws = get_or_create_sheet(wb, "Maflow Plastics Poland")
                        process_MaflowPlastic(text, filename, ws)

                    elif "ITW Slovakia s.r.o." in text:
                        ws = get_or_create_sheet(wb, "ITW Slovakia s.r.o.")
                        process_ITWSlovakiasro(text, filename, ws)

                    elif "Boryszew Kunststofftechnik" in text:
                        ws = get_or_create_sheet(wb, "Boryszew Kunststofftechnik")
                        process_BoryszewKunststofftechnik(text, filename, ws)

                    elif "Pro-X Automotive AG" in text:
                        if "CZ-571 01 MORAVSKA TREBOVA" in text:
                            ws = get_or_create_sheet(wb, "Pro-X MORAVSKA TREBOVA")
                            process_ProXMORAVSKATREBOVA(text, filename, ws)
                        elif "D-91555 FEUCHTWANGEN" in text:
                            ws = get_or_create_sheet(wb, "Pro-X FEUCHTWANGEN")
                            process_ProXMORAVSKATREBOVA(text, filename, ws)
                        elif "Eckerle Spritz" in text:
                            ws = get_or_create_sheet(wb, "Pro-X Eckerle Spritz")
                            process_ProXMORAVSKATREBOVA(text, filename, ws)

                    elif "Finke Anwendungstechnik GmbH" in text:
                        if "ABS XR 401 BK 9001" in text:
                            ws = get_or_create_sheet(wb, "Finke Anwendungstechnik_ABS XR 401 BK 9001")
                            process_FinkeAnwendungstechnik(text, filename, ws)
                        elif "High gloss ASA: LI941F" in text:
                            ws = get_or_create_sheet(wb, "Finke Anwendungstechnik_High gloss ASA: LI941F")
                            process_FinkeAnwendungstechnik(text, filename, ws)

                    elif "Formzeug GmbH" in text:
                        ws = get_or_create_sheet(wb, "Formzeug GmbH")
                        process_FormzeugGmbH(text, filename, ws)
                        
                    elif "ABC Technologies Karl Etzel GmbH" in text and text != full_text:
                        ws = get_or_create_sheet(wb, "ABC Technologies Karl Etzel GmbH")
                        process_ABCTechnologiesKEGmbH(full_text, filename, ws)       

                    elif "Plant Oldenburg" in text:
                        ws = get_or_create_sheet(wb, "Plant Oldenburg")
                        process_PlantOldenburg(text, filename, ws)      

                    elif "c/o Linden GmbH" in text:
                        if "30022062 ABS XR 410 NATUR" in text:
                            ws = get_or_create_sheet(wb, "CO Linden GmbH_30022062 ABS XR 410 NATUR")
                            process_coLindenGmbH(text, filename, ws)
                        elif "30021896 LG LI941V 94841" in text:
                            ws = get_or_create_sheet(wb, "CO Linden GmbH_30021896 LG LI941V 94841")
                            process_coLindenGmbH(text, filename, ws)

                    elif "Plant Neustadt" in text:
                        ws = get_or_create_sheet(wb, "Plant Neustadt")
                        process_PlantNeustadt(text, filename, ws)

                    elif "SMP Automotive Technology Ibérica SLU" in text:
                        ws = get_or_create_sheet(wb, "SMP Automotive Technology Ibérica SLU")
                        process_SMPAutomotiveTechnologyIbéricaSLU(text, filename, ws)

                    elif "SMR Automotive Mirror Technology" in text:
                        if "ASA LI941V" in text:
                            ws = get_or_create_sheet(wb, "SMR Hungary Bt_ASA LI941V")
                            process_SMRAutomotiveMirrorTechnology(text, filename, ws)
                        elif "LG LI941 F 94484 PIANO BLACK" in text:
                            ws = get_or_create_sheet(wb, "SMR Hungary Bt_LG LI941 F 94484 PIANO BLACK")
                            process_SMRAutomotiveMirrorTechnology(text, filename, ws)
                        elif "ASA LI941F-94841" in text:
                            ws = get_or_create_sheet(wb, "SMR Hungary Bt_ASA LI941F-94841")
                            process_SMRAutomotiveMirrorTechnology(text, filename, ws)

                    elif "Plant Schierling" in text:
                        ws = get_or_create_sheet(wb, "Plant Schierling")
                        process_PlantSchierling(text, filename, ws)
                            
                    elif "Uwe Etzel GmbH" in text:
                        ws = get_or_create_sheet(wb, "Uwe Etzel GmbH")
                        process_UweEtzelGmbH(text, filename, ws)

                    elif "SLM Kunststofftechnik GmbH" in text:
                        ws = get_or_create_sheet(wb, "SLM Kunststofftechnik GmbH")
                        process_SLMKunststofftechnikGmbH(text, filename, ws)

                    else:
                        print(f"⚠️ {filename}: 지정된 키워드 없음. 스킵.")
                        No_Key_Word.append(filename)
                        print(f"작업 진행 되어야 되는 파일들: {No_Key_Word}")

            except Exception as e:
                print(f"❌ {filename} 처리 중 오류 발생: {e}")

    for sheet in wb.sheetnames:
        ws = wb[sheet]

        max_row_f = ws.max_row
        for row in range(max_row_f, 0, -1):
            if ws[f'F{row}'].value is not None:
                last_row_f = row

                #######################################
                ## 값 못 받아오는 부분 때문에 임시 코드 ##
                #######################################
                if last_row_f < 6:
                    last_row_f = 6
                break

        for row in range(6, last_row_f + 1):
            F_value = f"F{row}"
            if F_value:
                ws[f"E{row}"] = f"=E{row-1}-F{row}"
        
        ## PO Number page가 달라 못받을 경우 H6에 있는 값 받아오기
        fill_missing_pno(ws)

        ## 색상 추가!! 필요 없을 경우 아래 한줄만 삭제
        apply_conditional_formatting(ws, last_row_f)  # 각 시트에 조건부 서식 적용

        ws.freeze_panes = 'B5'

    if os.path.exists(output_excel):
        os.remove(output_excel)
    wb.save(output_excel)
    print(f"✅ 함수 추가 및 엑셀 파일 저장 완료: {output_excel}")
       

# 실행
folder_path = r"C:\Users\82109\Desktop\개인\Python Test\마테리얼 코드 모름"
excel_path =r"C:\Users\82109\Desktop\개인\Python Test"
output_excel = os.path.join(excel_path, f'{datetime.now().strftime("%Y-%m-%d")}.xlsx')
datetime.now().strftime("%Y-%m-%d")

extract_info(folder_path, output_excel)
